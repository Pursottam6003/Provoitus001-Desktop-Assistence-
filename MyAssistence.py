#imports 
from email import message
from logging import exception
from urllib import response
import webbrowser
import pywhatkit
from xmlrpc.client import DateTime
import pyttsx3 #for text to speech 
import speech_recognition as sr   # for speech to text 
import pyaudio  # audio
import os # importing os 
import time 
import pyautogui 
import smtplib 
import openpyxl 
from time import gmtime,strftime
from datetime import date
import datetime
from playsound import playsound 
from threading import Thread
import requests,json 
import time 
import csv 

engine=pyttsx3.init()
voices =engine.getProperty('voices')
engine.setProperty('voice',voices[1].id)
rate=engine.setProperty('rate',150)
engine.setProperty('volume',0.9)


workbook = openpyxl.load_workbook('email_name.xlsx')

sheet= workbook.active
myfinaldict={} #declearing initial empty dictionary 

for i in range(1, 112):
    value1=sheet.cell(row=i,column=1).value
    value2 =sheet.cell(row=i,column=2).value
    mydict={value1.lower():value2.lower()}
    myfinaldict.update(mydict)


contact = open('contact.csv',encoding="utf8")

csvreader= csv.DictReader(contact)
mycontact={} #contacts 
for a in csvreader :
    value1=a['Name']
    value2=a['Numbers']
    mydict={value1.lower():value2.lower()}
    mycontact.update(mydict)

def say(message):
    engine.say(message)
    engine.runAndWait()
    engine.stop()

#defining and calling functions namely
# Listening the user and seding the  string that he got 
def takecomand():
    r=sr.Recognizer() 
    with sr.Microphone() as source:
        print('listening...')
        r.adjust_for_ambient_noise(source,duration=1)
        audio=r.listen(source)
    #recongnising under try catch block
    try :
        print('Recongnising...')
        query=r.recognize_google(audio,language='en-in') 
        query=query.lower()
        return query
    except Exception as e:
        print(e)
        print('Sorry did not get you')
        engine.say('sorry did not get you please try again')
        engine.runAndWait()
        engine.stop()
        return "none"


def call_somebody():
    say('To whome would you like to call')
    print('Please tell the name the person')

    friend_call=takecomand() 
    print(f'User saild: {friend_call}')
    friend_call=friend_call.replace('call','')
    
    print(f"calling {friend_call}")

    pyautogui.click(201,1061)
    pyautogui.typewrite('phone')    #for clicking on the taskbar 
    time.sleep(2)
    pyautogui.click(560,568)    #for selecting from the apps 
    time.sleep(1)
    pyautogui.click(1600,116)
    pyautogui.typewrite(friend_call)  # contact name of person 
    time.sleep(1)
    pyautogui.click(1597,173)   # selcting the topmost match
    time.sleep(1)
    say(f"would you like to call {friend_call} now please say yes or no ")
    controll=takecomand()

    if 'yes' in controll :
        pyautogui.click(1663,822)   # finally calling by clicking the call button 
    
    else :
        say('Ok sir now you can controll the comand manually thank you for using me')

     
         

def send_whats(number,message):
    # pywhatkit.sendwhatmsg(number,message,20,53)
    # time.sleep(5)
    # pyautogui.click(1774,962)
    hour = datetime.datetime.now().hour
    minutes= datetime.datetime.now().minute
    minutes=minutes+2
    print(hour)
    print(minutes)
    if minutes >=58 :
        if hour <=22:
            hour=hour+1
        
        else :
            hour=0
    
    pywhatkit.sendwhatmsg(number,message,hour,minutes)
    time.sleep(5)
    pyautogui.click(1774,962)
    say('messaged sucessfully')
    print('Message Sent')



def send_whatsappmsg():
    say('Would you like to message whome')
    print("Please tell whome to message in whatsapp")
    friend=takecomand()
    print(f"User said {friend}")

    if 'message' in friend :
        friend =friend.replace('message','')
        friend=friend.lstrip()
    num =mycontact.get(friend)
    # print(num)
    # print(friend)

    if 'none' in num :
        say('Sorry ! cannot find the contact please try again')
        print("Could'nt find the contact from contact list")
        return 
    
    else :
        
        print(f'Messaging to {friend}')
        say(f"Messaging to {friend} ")
        say(f'what would you like to message your friend {friend}')
        message=takecomand()
        print(f"Message is {message}")
        send_whats(num,message)
          

          

def send_mail(emailid,message):
    #creat SMTP session 
    s = smtplib.SMTP('smtp.gmail.com', 587)

    #start for TLS security 
    s.starttls()

    #authentication 
    s.login("rahulsah6003@gmail.com","rahul@12345")

    # message ="Hello my dear friend i am pursottam"

    # s.sendmail("rahulsah6003@gmail.com","pursottamsah@outlook.com",message)
    s.sendmail("rahulsah6003@gmail.com",emailid,message)
    print('Email sent sucessfully')
    s.quit()

def send_email():
    say('To whome would you like to send email')

    message=takecomand()
    print(message)
    for name in myfinaldict.keys():
        if name in message :
            emailid=myfinaldict[name]
            print(f"Sending mail to {name} with respective email Id{emailid}")
            say('What message would you like to send')
            print('What message would you like to send')
            descripion=takecomand()
            send_mail(emailid,descripion)


def weather():
    responses = requests.get('https://goweather.herokuapp.com/weather/Tinsukia')
    weather_data = json.loads(responses.text)
    description= weather_data['description']
    temperature=weather_data['temperature']
    windspeed=weather_data['wind']
    forcast=weather_data['forecast']
    #for 7 days we can use 
    print("Todays weather\n")
    print(f"Overall Weather : {description}")
    print(f"Temperature : {temperature}")
    print(f"Windspeed : {windspeed}")
    temperature =temperature.replace('°C',' ')
    windspeed1=windspeed.replace('km/h',' ')
    say(f"Todays weather {description}")
    say(f"Temperature is {temperature} degree celsius")
    say(f"Windspeed :{windspeed1} kilometers per hour" )
    

def asking_time():
    print('   Todays Date\t'+'  '+' Time')    #date for todays date decorating with strftime and localtime 
    curr_t=time.localtime()
    myTime=time.strftime("%H:%M:%S",curr_t)
    date= str(strftime("%a, %d %b %Y"))
    print("╔═.✵.════════════════════════╗")
    print(date,end='  ')
    print(myTime)
    print("╚════════════════════════.✵.═╝")
     
   
    say(date)
    say(myTime)
  

def tell():
    say('What would you like to open')
   

def news():
    url='https://saurav.tech/NewsAPI/top-headlines/category/general/in.json'
    response=requests.get('https://saurav.tech/NewsAPI/top-headlines/category/general/in.json')
    json_data=json.loads(response.text)

    content=json_data['articles']
    say('Here is your top 10 headlines of the day !')

    for i in range(10):
        print(content[i]['title'])
        say(content[i]['title'])
        print(content[i]['description'])
        say(content[i]['description'])
        print(content[i]['author'])
        say(f"This article was contributed by { content[i]['author']}")

def background_play():
    playsound("music3.mp3")

def getquotes():
    response = requests.get('https://api.quotable.io/random?tags=motivation|inspirational')
    json_data= json.loads(response.text)
    quote = json_data['content']
    author = json_data['author']
    return quote,author

def play_bhajans():
    link='https://wynk.in/music/playlist/bhajans/bb_1469093087164'
    webbrowser.open_new_tab(link)

def work():
    #now we are creating the application that used in our daily work 
    tell()
 
    task=takecomand()

    print('getting results...')

    print(f"user said : {task}")

    if 'youtube' in task :
        
        webbrowser.open_new_tab('https://www.youtube.com/')
        return "none"
    
    elif 'linkdin' in task:
        webbrowser.open_new_tab('https://www.linkedin.com/')
        return "none"
    
    elif 'moodle ' in task:
        webbrowser.open_new_tab('http://nitap.co.in/moodle/')
        return "none"

    elif 'facebook' in task :
        
        webbrowser.open_new_tab('https://www.facebook.com/')
        return "none"
    
    elif 'google meet' in task:
        #may ask the code also 
        
        webbrowser.open_new_tab('https://meet.google.com/landing?authuser=1')
        return "none" 
    
    elif 'whatsapp' in task:
        
        webbrowser.open_new_tab('https://web.whatsapp.com/')
        return "none"
    
    elif 'webex' in task:
        webbrowser.open_new_tab('https://nitap.webex.com/wbxmjs/joinservice/sites/nitap/meeting/download/3e0f81c74f544a6d9ebc1b2a60d47e68?siteurl=nitap&MTID=mfaad87feb04e2f95b3f227d5ed7dbc58')
        return "none"
    elif 'zoom' in task:
        
        webbrowser.open_new_tab('https://zoom.us/join')
        return "none"
    
    elif 'jitsi' in task:
        
        webbrowser.open_new_tab('https://meet.jit.si/')
        return "none"

    else :
        link ="https://www.google.com.tr/search?q={}".format(task)
        webbrowser.open_new_tab(link)
        return "none"

def play_games():
    say('What type of game would you like to play')
    say('action puzzle shooting sports games defence games, driving game ,strategy game ,arcade game adventure')
    print('1. Action Game \t 2.Shooting Game \t3.Sports\t 4.Defence Game\t 5.Driving Game\t 6.Strategy Game \t7. Aracade Games\t 8.Adventurous Games')
    game=takecomand()

    part='''  
    __   _(_) __| | ___  ___     __ _  __ _ _ __ ___   ___  ___ 
    \ \ / / |/ _` |/ _ \/ _ \   / _` |/ _` | '_ ` _ \ / _ \/ __|
     \ V /| | (_| |  __/ (_) | | (_| | (_| | | | | | |  __/\__ \\
      \_/ |_|\__,_|\___|\___/   \__, |\__,_|_| |_| |_|\___||___/
                                |___/                           
    '''                                              
 
    print(part)
    
    if 'action' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/action/')
    
    elif 'puzzle' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/puzzles/')
    
    elif 'shooting' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/shooting/')
    
    elif 'sports' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/sports/')

    elif 'defense' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/defense/')

    elif 'driving' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/driving/')

    elif 'board games' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/board/')

    elif 'adventure' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/adventure/')

    elif 'arcade' in game:
        webbrowser.open_new_tab('https://miminogames.com/games/arcade/')

    else :
        url = "https://www.google.com.tr/search?q={}".format(game)
        webbrowser.open(url)
        return "none"


def open_something():
    tell()
    print("What whould you like to open !")
    query=takecomand()
    if 'work' in query :
        work()
        return "None"


    elif 'computer' in query:
        
        say('To whome would you like to Open : 1. Microsoft Visual Studio code 2. Android Studios 3. Microsoft word 4.Calculator 5.Microsoft Powerpoint 6.office 7.Microsoft excel 8.Team viewer 9.Settings 10. Alrams and clock 11.phone 12.One note')

        print('To whome would you like to Open : 1. Microsoft Visual Studio code \t2. Android Studios \t3. Microsoft word\t 4.Calculator\t 5.Microsoft Powerpoint \t6.office \t7.Microsoft excel\t 8.Team viewer \t9.Settings \t10. Alrams and clock \t11.phone \t12.One note\t')

        app=takecomand() 
        print(f'User saild: Open {app}')
        
        if 'open' in app :
            app=app.replace('open','')
        
    
        pyautogui.click(201,1061)
        pyautogui.typewrite(app)    #for clicking on the taskbar 
        time.sleep(2)
        pyautogui.click(560,568)    #for selecting from the apps 
        time.sleep(1)
        say(f'The {app} opened sucessfully thank you')
        print(f'The {app} opened sucessfully thank you')
     
        
        
def intro():
    visual_efftect()
    print('Hello Sir...')
    greet()
    say('Hello Mr CEO ! Sir I am ProBot-001 , your virtual personal assistent \n')
    say('Here I am happy to help you please tell me how i can help you ?\n')
    say('Please look on the screen  some commands and tips i can perform happily')
    print('\n\n\n')

    print('Would you like to select the Assistent Controll\n 1 for voice controll or 2 for  keyboard')
    control=int(input())

    if control == 1 :
        print('Ask me anything')
        querry=takecomand()#we will update the voice settings soon 
        
    elif control == 2 :
        print('Please tell Me how i can help you\n')
        print('1.Asking TIme \t 2. Greeting \t 3. Daily Task \t 4.Chanting Bhajans\t 5. Reading News\t 6.Opening Social Media')
        print('7. Weather Report \t 8.Motivate Me 9.Send Email to friend  \t 10.Sending Messages through whatsapp\t 11.Play games  \t12.Call Your friend\t')
         
        want_choice=int(input('Select choice'))

        if want_choice == 1 :
            asking_time()
        
        elif want_choice==2:
            greet()
        
        elif want_choice==3:    #opening something 
            open_something()
            
        elif want_choice==4:
            play_bhajans()    #play bhajans 

        elif want_choice==5:
            news()    #read news 

        elif want_choice==6:
            open_something()    #opening social media 

        elif want_choice==7:
            weather()   # check weather

        elif want_choice==8:
            getquotes()    #mptivate me  

        elif want_choice==9:
            send_email()
            
        elif want_choice==10:
            send_whatsappmsg()
        
        elif want_choice ==11:
            play_games()
        
        elif want_choice == 12 :
            call_somebody()
        else :
            pass 
    else :
        print('Please Select a valid input')
        return 
  
    #just to wait 



def visual_efftect():
    
    print('\t\t\t\t\t\t\t  »»-————————　★ 　————————-««')
    time.sleep(0.3)
    print('\t\t\t\t\t\t\t    Starting the assistent')
    time.sleep(0.3)
    print('\t\t\t\t\t\t\t  »»-————————　★ 　————————-««')
 
    background_play()   #background sound that will work 
    print('\n\n')
    time.sleep(0.4)
    for i in range(20):
        print("🔊 ",end=' ✧ ✧ ')
         
    #print hello sir 
    print('\n\n')
    multiline_str=''' 
                                                    *✲*´*。.❄¯*✲。❄。*¨`*✲´*。❄¨`*✲。❄。*`*
                                                    *╔════════════ ༺❀༻❤༺❀༻ ════════════╗*
                                                    *♥*❄¯*✲❄♫♪♩░H░E░L░L░O░░S░I░R░ ♫♫♪❄¯*✲❄
                                                    *╚════════════ ༺❀༻❤༺❀༻ ════════════╝*
                                                    *✲*´*。.❄¯*✲。❄。*¨`*✲´*。❄¨`*✲。❄。*`*
    '''
    print(multiline_str)
    
def greet():
    print("╔═.✵.════════════════════════╗")
    print('   Todays Date\t'+'  '+' Time')    #date for todays date decorating with strftime and localtime 
    print("╚════════════════════════.✵.═╝")
    curr_t=time.localtime()
    myTime=time.strftime("%H:%M:%S",curr_t)
    date= str(strftime("%a, %d %b %Y"))
    print('--------------------------')
    print(date,end='  ')
    print(myTime)
    hours = datetime.datetime.now().hour
    print('--------------------------')
    quote,author=getquotes()

    say('Your Quote of The Day given by'+author)
    say(quote)
    child_happyr=''' 
            (✿ ͡◕ ᴗ◕)つ━━✫・*。
            ⊂　　 ノ 　　　・゜+.
            しーーＪ　　　°。+ *´¨)
            .· ´𝔩Quote of The Day ☆´¨) ¸.·*¨)
            (¸.·´ (¸.·’* (¸.·’* (¸.·’* (¸.·’* (¸.·’* *¨)
        '''
    print(child_happyr)
    print('\n\n\n')
    print('\t\t\t\t\t\t\t  »»-————————　★ 　————————-««')
    print('\t\t\t\t\t\t\t     Your Quote of The Day ')
    print('\t\t\t\t\t\t\t  »»-————————　★ 　————————-««')
 
    print('<====  '+quote+'  ====>')
    print('\t\t\t\t\t\t\t\t *:･ﾟ✧*:･ﾟ✧ ---'+author)
    if hours <1:
        print('Good Hardworking Sir ! But its time to relax and sleep')
        say('Ohh ! Sir please sleep tommrow you have meeting')
        say('Good Night Sir have sweet dreams !')
    
    elif hours >6 and hours <12 :
        say('Good Moring Sir')
        say('Here is your morning quotes from my side !')
        print('One small positive thought can change your whole day')
        engine.say('One small positive thought can change your whole day')

    elif hours >12 and hours <=15:
        say('Good Afternoon Sir')
    elif hours >15 and hours <20 :
        say('Good Evening Sir')
    else :
        say('Keep Growing Sir! keep Going !')

if __name__=="__main__" :
    intro()

    
 
